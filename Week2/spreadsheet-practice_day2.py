# You will continue to work on the inventory spreadsheet that you created from yesterday's exercise
# import the appropriate method from the correct module
from openpyxl import load_workbook

# Import the workbook that you created in yesterday's exercise from
wb = load_workbook("Week2\spreadsheets\inventory.xlsx")

# verify what sheet names are available
for sheet in wb:
    print(sheet.title)

# access and store the appropriate worksheet to the variable 'ws'
ws = wb['CURRENT_MONTH_INVENTORY']

# Print out the cell values for each row
all_values = ws.values
print(list(all_values))

# Create a column header within that worksheet called order_amount
ws["F1"] = "order_amount"

# save the latest changes
wb.save("Week2\spreadsheets\inventory.xlsx")

wb.close()