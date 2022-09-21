# import the appropriate method from the correct module
from openpyxl import load_workbook

# Import the workbook that you updated in today's practice from
# "./spreadsheets/inventory.xlsx"
wb = load_workbook("Week2\spreadsheets\inventory.xlsx")

# access and store the appropriate worksheet to the variable 'ws'
ws = wb.active
print(ws.title)

# Define a function called add_order_amount that takes in a single parameter called 'row'
def add_order_amount(row):
    # TIP: create variables for quantity, threshold, max_amount that retrieves the values first for cleanliness
    max_amount = ws.cell(row=row, column=3).value
    threshold = ws.cell(row=row, column=4).value
    quantity = ws.cell(row=row, column=5).value
    # IF the quantity is less than or equal to the threshold,
    # THEN set the order_amount to the difference between the threshold and the quantity
    # than calculate the order_amount (max_amount - quantity) 
    # assign the value to that row, column 6
    if quantity <= threshold:
        order_amount = max_amount - quantity
        ws.cell(row=row, column=6).value = order_amount

# Perform a for..in loop through the range(2, len(ws.rows))
#   - call the function add_order_amount() passing in the number of the range
for num in range(2, 15):
    add_order_amount(num)

wb.save("Week2\spreadsheets\inventory.xlsx")

wb.close()