# A1. from the appropriate library, import only the Workbook
from openpyxl import Workbook

# A2. Before anything, we need a workbook to work with.. (lecture line # 27)
wb = Workbook()

# A3. We need to interact with a single worksheet. (lecture line # 33)
ws = wb.active

# A4. assign the value of "First Name" to A1
ws['A1'] = "First name"

# A5. assign the value of "Last Name" to B1
ws['B1'] = "Last name"

# STOP HERE - RETURN TO LECTURE

# B1. For all of column A, starting at row 2 until row 10, make the cell values: "Gabriel" (attempt a loop)
colA = ws['A']
for cell in range(2,11):
    ws["A" + str(cell)] = "Gabriel"

last_names = ['Rolley', 'Smith', 'Balenga', 'Issac', 'Cruise', 'Depp', 'Heard', 'Qiao', 'Biden']

# B2. Loop through a range from row 2 to 10 and assign the cell value to last names according to index in column B
# NOTE: PAY ATTENTION to the starting number of the range and how it differs from the starting index of the list
for cell in range(2,11):
    ws["B" + str(cell)] = last_names[cell - 2]
    # print(ws["B" + str(cell)].value)
        

# B3. Save the file
wb.save("Week2\spreadsheets\day_1_practice.xlsx")

wb.close()