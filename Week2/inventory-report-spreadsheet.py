# SCENARIO: You're a store owner receiving the inventory report for this month.
# You will receive the product order spreadsheet soon and it is easier to calculate your order
# if your inventory was also on a spreadsheet! Recreate the following spreadsheet with Python: 

# don't forget your appropriate imports.
# assign the title of the initial active sheet to "CURRENT_MONTH_INVENTORY"

#   product_name    product_id  max_amount      reorder_threshold   quantity
# 	oreo            2323        1000            300                 743
# 	coke            6545        500             100                 101
# 	pepsi	        3456        200             50                  137
# 	lays_chip	    4567        1500            500                 364
# 	pringles	    2134        2000            600                 120
# 	sour_worms	    2362        100             10                  85
# 	choco_cookies   0923	    200             25                  24
# 	donuts	        2786        200             25                  12
# 	hot_dogs	    6723        100             10                  39
# 	ice_cream	    9237        200             50                  234
# 	gum	            2092        3500            1000                1232
# 	pretzels        8246	    100             5                   11
# 	kit_kat	        9276        1000            250                 249

# TIP: create a list of each column (ie. product_names = [oreo, ...]) and use those to loop through :)
from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws.title = "CURRENT_MONTH_INVENTORY"
# print(ws.title)

column_titles = ["product_name", "product_id", "max_amount", "reorder_threshold", "quantity"]
prod_titles = ["oreo", "coke", "pepsi", "lays_chip", "pringles", "sour_worms", "choco_cookies", "donuts", "hot_dogs", "ice_cream", "gum", "pretzels", "kit_kat"]
prod_ids = ["2323", "6545", "3456", "4567", "2134", "2362", "0923", "2786", "6723", "9237", "2092", "8246", "9276"]
max_amt = [1000, 500, 200, 1500, 2000, 100, 200, 200, 100, 200, 3500, 100, 1000]
reorder_thresh = [300, 100, 50, 500, 600, 10, 25, 25, 10, 50, 1000, 5, 250]
quantity_num = [743, 101, 137, 364, 120, 85, 24, 12, 39, 234, 1232, 11, 249]

for colTitles in range(1, len(column_titles) + 1):
    ws.cell(row=1, column=colTitles).value = column_titles[colTitles - 1]
    # print(column_titles[colTitles-1])

for prod in range(2, len(prod_titles) + 2):
    ws.cell(row=prod, column=1).value = prod_titles[prod - 2]
    
for id in range(2, len(prod_ids) + 2):
    ws.cell(row=id, column=2).value = prod_ids[id - 2]
    
for amt in range(2, len(max_amt) + 2):
    ws.cell(row=amt, column=3).value = max_amt[amt - 2]
    
for thres in range(2, len(reorder_thresh) + 2):
    ws.cell(row=thres, column=4).value = reorder_thresh[thres - 2]
    
for qual in range(2, len(quantity_num) + 2):
    ws.cell(row=qual, column=5).value = quantity_num[qual - 2]

# save your file
wb.save("Week2\spreadsheets\inventory-spreadsheet.xlsx")

wb.close()